from __future__ import annotations

import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook

from .dates import format_pickup_date, format_warranty_date
from .model_names import normalize_model_name
from .models import CustomerProfile, IntakeRecord, WarrantyInfo


NETSUITE_HEADERS = [
    "Internal ID",
    "Company",
    "Service Board",
    "Asset ID",
    "Serial Number",
    "Customer Name",
    "Subject",
    "Assigned To",
    "Contact",
    "In Facility Location",
    "Device Type",
    "PC Information",
    "Pickup Date",
    "Bin #",
    "Warranty Expiration",
    "Repair status",
    "Repair type",
    "Reason",
]

DEFAULT_PROFILE_DATA = {
    "humble": {
        "company": "1006 GTS Internal : GTS - Break/Fix - OEM Reimbursements",
        "service_board": "GTS Break/Fix",
        "customer_name": "  Humble ISD",
        "contact": "434 GTS Internal : 1006 GTS - Break/Fix - OEM Reimbursements : Humble ISD Break/Fix",
        "in_facility_location": "North Gessner : Humble ISD (BF)",
        "device_type": "Laptop",
    }
}


def load_customer_profiles(path: str | Path) -> dict[str, CustomerProfile]:
    path = Path(path)
    if path.exists() and path.suffix.lower() in {".xlsx", ".xlsm"}:
        profiles = _load_profiles_from_template_workbook(path)
        if profiles:
            return profiles
        raise ValueError(f"No customer profile rows were found in '{path}'.")

    data = json.loads(path.read_text(encoding="utf-8")) if path.exists() else DEFAULT_PROFILE_DATA
    profiles: dict[str, CustomerProfile] = {}
    for key, values in data.items():
        profile = CustomerProfile(
            key=key,
            company=values["company"],
            service_board=values["service_board"],
            customer_name=values["customer_name"],
            contact=values["contact"],
            in_facility_location=values["in_facility_location"],
            device_type=values.get("device_type", "Laptop"),
            assigned_to=values.get("assigned_to", ""),
        )
        profiles[key] = profile
    return profiles


def build_netsuite_row(
    record: IntakeRecord,
    profile: CustomerProfile,
    warranty: WarrantyInfo | None = None,
    pickup_date_override: object = None,
    model_map: dict[str, str] | None = None,
    include_assigned_to: bool = False,
) -> dict[str, str]:
    warranty = warranty or WarrantyInfo(serial=record.serial)
    model = normalize_model_name(record.model, model_map) or warranty.model
    pickup_date = pickup_date_override if pickup_date_override not in (None, "") else record.pickup_date
    warranty_end = format_warranty_date(record.warranty_expiration) or warranty.warranty_end

    return {
        "Internal ID": "",
        "Company": profile.company,
        "Service Board": profile.service_board,
        "Asset ID": record.asset_id,
        "Serial Number": record.serial,
        "Customer Name": profile.customer_name,
        "Subject": f"{record.asset_id} {record.serial} {profile.customer_name}",
        "Assigned To": profile.assigned_to if include_assigned_to else "",
        "Contact": profile.contact,
        "In Facility Location": profile.in_facility_location,
        "Device Type": profile.device_type,
        "PC Information": model,
        "Pickup Date": format_pickup_date(pickup_date),
        "Bin #": record.bin_location,
        "Warranty Expiration": warranty_end,
        "Repair status": "",
        "Repair type": "",
        "Reason": "",
    }


def write_netsuite_csv(path: str | Path, rows: list[dict[str, str]], pad_rows: int = 0) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=NETSUITE_HEADERS, lineterminator="\r\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
        for _ in range(max(0, pad_rows - len(rows))):
            writer.writerow({header: "" for header in NETSUITE_HEADERS})


def _load_profiles_from_template_workbook(path: Path) -> dict[str, CustomerProfile]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = _find_sheet(workbook, "Template Data")
        if sheet is None:
            raise ValueError(f"Sheet 'Template Data' was not found in '{path.name}'. Available sheets: {', '.join(workbook.sheetnames)}")

        headers = [_normalize_header(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        columns = {header: index for index, header in enumerate(headers) if header}
        required = ["company", "serviceboard", "customername", "contact", "infacilitylocation"]
        missing = [header for header in required if header not in columns]
        if missing:
            raise ValueError(f"Template Data is missing required column(s): {', '.join(missing)}")

        profiles: dict[str, CustomerProfile] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            customer_name = _cell(row, columns, "customername")
            contact = _cell(row, columns, "contact")
            in_facility_location = _cell(row, columns, "infacilitylocation")
            if not (customer_name and contact and in_facility_location):
                continue
            profile = CustomerProfile(
                key=_profile_keys(customer_name)[0],
                company=_cell(row, columns, "company"),
                service_board=_cell(row, columns, "serviceboard"),
                customer_name=customer_name,
                contact=contact,
                in_facility_location=in_facility_location,
                device_type=_cell(row, columns, "devicetype") or "Laptop",
                assigned_to=_cell(row, columns, "assignedto"),
            )
            for key in _profile_keys(customer_name):
                profiles.setdefault(key, profile)
        return profiles
    finally:
        workbook.close()


def _find_sheet(workbook, sheet_name: str):
    for candidate in workbook.worksheets:
        if candidate.title.lower() == sheet_name.lower():
            return candidate
    return None


def _normalize_header(value: object) -> str:
    if value in (None, ""):
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def _cell(row: tuple[object, ...], columns: dict[str, int], header: str) -> str:
    index = columns.get(header)
    if index is None or index >= len(row):
        return ""
    value = row[index]
    return "" if value in (None, "") else str(value).strip() if header != "customername" else str(value)


def _profile_keys(customer_name: str) -> list[str]:
    clean = customer_name.strip().lower()
    clean = clean.replace("b/f", "bf")
    normalized = re.sub(r"[^a-z0-9]+", "_", clean).strip("_")
    keys = [normalized] if normalized else []
    if normalized.endswith("_isd"):
        keys.append(normalized.removesuffix("_isd"))
    if normalized.endswith("_bf"):
        keys.append(normalized.removesuffix("_bf"))
    return list(dict.fromkeys(keys))
