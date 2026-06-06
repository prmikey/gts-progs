from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .dates import date_sort_key, format_warranty_date
from .model_names import normalize_model_name
from .models import IntakeRecord, WarrantyInfo


FIELD_ALIASES = {
    "asset_id": {
        "assettag",
        "assetid",
        "asset",
        "tag",
    },
    "serial": {
        "serial",
        "serialnumber",
        "serialno",
        "serialimei",
        "serialorimei",
        "sn",
        "serial#",
    },
    "bin_location": {
        "location",
        "binloc",
        "binlocation",
        "bin#",
        "loc",
    },
    "pickup_date": {
        "pickupdate",
        "pickupdate",
        "datereceived",
        "daterecieved",
        "datereceieved",
        "datereceipted",
        "datein",
    },
    "warranty_expiration": {
        "warranty",
        "warrantyexpiration",
        "warrantyexpiry",
        "warrantyend",
    },
    "model": {
        "model",
        "devicemodel",
        "pcinformation",
        "pcinfo",
    },
    "machine_type": {
        "machinetype",
        "type",
    },
    "code": {
        "code",
        "entitlementcode",
    },
    "end": {
        "end",
        "enddate",
        "warrantyend",
        "warrantyexpiration",
    },
    "subseries": {
        "subseries",
        "product",
        "model",
    },
}


def normalize_header(value: object) -> str:
    if value in (None, ""):
        return ""
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum() or ch == "#")


def _field_for_header(header: object, wanted_fields: Iterable[str]) -> str | None:
    normalized = normalize_header(header)
    for field in wanted_fields:
        if normalized in FIELD_ALIASES[field]:
            return field
    return None


def _open_sheet(path: Path, sheet_name: str | None):
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name:
        for candidate in workbook.sheetnames:
            if candidate.lower() == sheet_name.lower():
                return workbook, workbook[candidate]
        raise ValueError(f"Sheet '{sheet_name}' was not found in {path.name}. Available sheets: {', '.join(workbook.sheetnames)}")
    return workbook, workbook[workbook.sheetnames[0]]


def _find_header_row(ws, wanted_fields: Iterable[str]) -> tuple[int, dict[str, int]]:
    wanted_fields = list(wanted_fields)
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True), start=1):
        found: dict[str, int] = {}
        for idx, value in enumerate(row):
            field = _field_for_header(value, wanted_fields)
            if field and field not in found:
                found[field] = idx
        if "serial" in found and (("asset_id" in found) or ("machine_type" in found)):
            return row_number, found
        if "serial" in found and {"code", "end"}.issubset(found):
            return row_number, found
    raise ValueError(f"Could not find a recognizable header row in sheet '{ws.title}'.")


def _clean_string(value: object, upper: bool = False) -> str:
    if value in (None, ""):
        return ""
    text = str(value).strip()
    return text.upper() if upper else text


def read_intake_workbook(path: str | Path, sheet_name: str = "Inbound_Deployment") -> list[IntakeRecord]:
    path = Path(path)
    workbook, ws = _open_sheet(path, sheet_name)
    try:
        header_row, columns = _find_header_row(
            ws,
            ["asset_id", "serial", "bin_location", "pickup_date", "warranty_expiration", "model"],
        )
        records: list[IntakeRecord] = []
        for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            asset_id = _clean_string(row[columns["asset_id"]]) if "asset_id" in columns and columns["asset_id"] < len(row) else ""
            serial = _clean_string(row[columns["serial"]]) if "serial" in columns and columns["serial"] < len(row) else ""
            if not asset_id and not serial:
                continue
            records.append(
                IntakeRecord(
                    asset_id=asset_id,
                    serial=serial,
                    bin_location=_clean_string(row[columns["bin_location"]]) if "bin_location" in columns and columns["bin_location"] < len(row) else "",
                    pickup_date=row[columns["pickup_date"]] if "pickup_date" in columns and columns["pickup_date"] < len(row) else None,
                    warranty_expiration=row[columns["warranty_expiration"]] if "warranty_expiration" in columns and columns["warranty_expiration"] < len(row) else None,
                    model=_clean_string(row[columns["model"]]) if "model" in columns and columns["model"] < len(row) else "",
                    row_number=row_number,
                )
            )
        return records
    finally:
        workbook.close()


def read_batch_template(path: str | Path, sheet_name: str | None = None) -> dict[str, str]:
    path = Path(path)
    workbook, ws = _open_sheet(path, sheet_name)
    try:
        header_row, columns = _find_header_row(ws, ["machine_type", "serial"])
        machine_types: dict[str, str] = {}
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            serial = _clean_string(row[columns["serial"]], upper=True) if columns["serial"] < len(row) else ""
            machine_type = _clean_string(row[columns["machine_type"]], upper=True) if columns["machine_type"] < len(row) else ""
            if serial and machine_type:
                machine_types[serial] = machine_type
        return machine_types
    finally:
        workbook.close()


def read_query_result(
    path: str | Path,
    sheet_name: str | None = "Result",
    preferred_code: str = "UWP",
    model_map: dict[str, str] | None = None,
) -> tuple[dict[str, WarrantyInfo], list[str]]:
    path = Path(path)
    workbook, ws = _open_sheet(path, sheet_name)
    warnings: list[str] = []
    try:
        header_row, columns = _find_header_row(ws, ["serial", "subseries", "code", "end"])
        grouped: dict[str, list[dict[str, object]]] = {}
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            serial = _clean_string(row[columns["serial"]], upper=True) if columns["serial"] < len(row) else ""
            if not serial:
                continue
            grouped.setdefault(serial, []).append(
                {
                    "code": _clean_string(row[columns["code"]], upper=True) if "code" in columns and columns["code"] < len(row) else "",
                    "end": row[columns["end"]] if "end" in columns and columns["end"] < len(row) else None,
                    "subseries": _clean_string(row[columns["subseries"]]) if "subseries" in columns and columns["subseries"] < len(row) else "",
                }
            )

        result: dict[str, WarrantyInfo] = {}
        for serial, rows in grouped.items():
            preferred = [row for row in rows if row["code"] == preferred_code.upper() and row["end"] not in (None, "")]
            candidates = preferred or [row for row in rows if row["end"] not in (None, "")]
            if not candidates:
                warnings.append(f"{serial}: query result has no warranty end date.")
                continue
            if not preferred:
                warnings.append(f"{serial}: no {preferred_code.upper()} row found; used latest available end date.")
            best = max(candidates, key=lambda row: date_sort_key(row["end"]))
            subseries = str(best.get("subseries") or "")
            result[serial] = WarrantyInfo(
                serial=serial,
                warranty_end=format_warranty_date(best["end"]),
                model=normalize_model_name(subseries, model_map),
                subseries=subseries,
                source=f"query:{best.get('code') or 'unknown'}",
            )
        return result, warnings
    finally:
        workbook.close()
